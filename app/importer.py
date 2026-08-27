"""Importador masivo de parcelas: la cooperativa que YA tiene la georreferencia de sus asociados
sube un Excel/CSV/GeoJSON y obtiene los mismos lotes → misma verificación → mismo dossier → mismo
sello blockchain que la captura en campo. El sello no depende de cómo entró el dato.

Encabezados flexibles (ES/EN). Una fila = una parcela. Las filas se agrupan en lotes por la columna
`lote` (si existe) o por productor. Polígonos: columna `poligono` = "lat,lon; lat,lon; ..." (>=3)."""
import csv, io, json, re, uuid, datetime
from .models import Lot, Plot, GeoPoint
from . import storage

MAX_ROWS = 1000

_H = {
    "productor": "producer", "producer": "producer", "nombre": "producer", "asociado": "producer",
    "producto": "commodity", "commodity": "commodity", "cultivo": "commodity",
    "region": "region", "región": "region", "zona": "region", "provincia": "region",
    "lote": "lot_key", "lot": "lot_key", "lot_id": "lot_key", "codigo_lote": "lot_key",
    "parcela": "plot_id", "plot": "plot_id", "plot_id": "plot_id", "codigo": "plot_id", "código": "plot_id",
    "lat": "lat", "latitud": "lat", "latitude": "lat", "y": "lat",
    "lon": "lon", "lng": "lon", "longitud": "lon", "longitude": "lon", "x": "lon",
    "area": "area_ha", "area_ha": "area_ha", "hectareas": "area_ha", "hectáreas": "area_ha", "ha": "area_ha",
    "cantidad": "quantity", "cantidad_kg": "quantity", "kg": "quantity", "quantity": "quantity", "volumen": "quantity",
    "poligono": "polygon", "polígono": "polygon", "polygon": "polygon",
    "vertices": "polygon", "vértices": "polygon", "coordenadas": "polygon", "wkt": "polygon",
}

_COMM = {"cafe": "coffee", "café": "coffee", "coffee": "coffee",
         "cacao": "cocoa", "cocoa": "cocoa", "chocolate": "cocoa"}


def _norm_header(h):
    return _H.get(str(h or "").strip().lower().replace(" ", "_"))


def _f(v):
    try:
        s = str(v).strip().replace(",", ".")
        return float(re.search(r"-?[\d.]+", s).group()) if s else None
    except Exception:
        return None


def _parse_polygon(txt):
    """'lat,lon; lat,lon; ...' (o 'lat lon;...' o WKT simple) → [GeoPoint]. None si no da >=3 puntos."""
    if not txt:
        return None
    nums = re.findall(r"-?\d+\.?\d*", str(txt))
    if len(nums) < 6:
        return None
    pts = []
    for i in range(0, len(nums) - 1, 2):
        a, b = float(nums[i]), float(nums[i + 1])
        # heurística: la latitud siempre está en [-90,90]; si el primer valor no puede serlo, es lon,lat (WKT)
        lat, lon = (a, b) if abs(a) <= 90 and abs(b) > 90 else ((a, b) if abs(a) <= 90 else (b, a))
        pts.append(GeoPoint(lat, lon))
    return pts if len(pts) >= 3 else None


def _rows_from_csv(data: bytes):
    text = data.decode("utf-8-sig", errors="replace")
    delim = ";" if text.count(";") > text.count(",") else ","
    return [r for r in csv.reader(io.StringIO(text), delimiter=delim)]


def _rows_from_xlsx(data: bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    return [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]


def _rows_from_geojson(data: bytes):
    """GeoJSON FeatureCollection → filas normalizadas (dicts)."""
    gj = json.loads(data.decode("utf-8", errors="replace"))
    out = []
    for f in (gj.get("features") or [])[:MAX_ROWS]:
        props = {str(k).lower(): v for k, v in (f.get("properties") or {}).items()}
        geom = f.get("geometry") or {}
        row = {"producer": props.get("producername") or props.get("productor") or props.get("producer") or "",
               "region": props.get("productionplace") or props.get("region") or "",
               "area_ha": _f(props.get("area")), "quantity": str(props.get("cantidad") or props.get("quantity") or ""),
               "commodity": str(props.get("producto") or props.get("commodity") or ""),
               "lot_key": str(props.get("lote") or props.get("lot_id") or ""),
               "plot_id": str(props.get("parcela") or props.get("plot_id") or ""), "lat": None, "lon": None, "polygon": None}
        if geom.get("type") == "Point":
            c = geom.get("coordinates") or []
            if len(c) >= 2:
                row["lon"], row["lat"] = float(c[0]), float(c[1])
        elif geom.get("type") == "Polygon":
            ring = (geom.get("coordinates") or [[]])[0]
            row["polygon"] = [GeoPoint(float(pt[1]), float(pt[0])) for pt in ring if len(pt) >= 2]
        out.append(row)
    return out


def parse(data: bytes, filename: str):
    """Devuelve (filas normalizadas [dict], errores [str])."""
    name = (filename or "").lower()
    if name.endswith(".geojson") or name.endswith(".json"):
        try:
            return _rows_from_geojson(data), []
        except Exception as e:
            return [], [f"GeoJSON inválido: {e}"]
    raw = _rows_from_xlsx(data) if name.endswith((".xlsx", ".xlsm")) else _rows_from_csv(data)
    if not raw:
        return [], ["Archivo vacío"]
    header = [_norm_header(h) for h in raw[0]]
    if "lat" not in header and "polygon" not in header:
        return [], ["No encuentro columnas de coordenadas. Usa la plantilla: productor, producto, region, "
                    "lat, lon, area_ha, cantidad_kg (opcional: lote, parcela, poligono)."]
    rows, errs = [], []
    for i, r in enumerate(raw[1:MAX_ROWS + 1], start=2):
        d = {}
        for j, key in enumerate(header):
            if key and j < len(r):
                d[key] = r[j]
        row = {"producer": str(d.get("producer") or "").strip(),
               "commodity": str(d.get("commodity") or "").strip(),
               "region": str(d.get("region") or "").strip(),
               "lot_key": str(d.get("lot_key") or "").strip(),
               "plot_id": str(d.get("plot_id") or "").strip(),
               "lat": _f(d.get("lat")), "lon": _f(d.get("lon")),
               "area_ha": _f(d.get("area_ha")),
               "quantity": str(d.get("quantity") or "").strip(),
               "polygon": _parse_polygon(d.get("polygon"))}
        if not row["polygon"] and (row["lat"] is None or row["lon"] is None):
            errs.append(f"Fila {i}: sin coordenadas"); continue
        if row["lat"] is not None and not (-90 <= row["lat"] <= 90 and -180 <= row["lon"] <= 180):
            errs.append(f"Fila {i}: lat/lon fuera de rango"); continue
        rows.append(row)
    if len(raw) - 1 > MAX_ROWS:
        errs.append(f"Se procesaron las primeras {MAX_ROWS} filas (límite por carga).")
    return rows, errs


def create_lots(rows, coop_id, coop_name, captured_by):
    """Agrupa filas en lotes (por `lote` o por productor) y los guarda. Devuelve resumen."""
    now = datetime.datetime.utcnow().isoformat() + "Z"
    groups = {}
    for r in rows:
        key = r["lot_key"] or r["producer"] or "sin-nombre"
        groups.setdefault(key, []).append(r)
    created = []
    for key, rs in groups.items():
        first = rs[0]
        commodity = _COMM.get(first["commodity"].lower(), first["commodity"].lower() or "coffee")
        plots, total_kg = [], 0.0
        for n, r in enumerate(rs, start=1):
            pid = r["plot_id"] or f"P{n}"
            pts = r["polygon"] if r["polygon"] else [GeoPoint(r["lat"], r["lon"])]
            plots.append(Plot(pid, pts, r["area_ha"]))
            q = _f(r["quantity"])
            if q:
                total_kg += q
        lot_id = "LOT-" + uuid.uuid4().hex[:8].upper()
        lot = Lot(lot_id, first["producer"] or key, coop_name, commodity, "PE",
                  first["region"], plots, "", "", (f"{total_kg:.0f} kg" if total_kg else ""),
                  coop_id, captured_by, now, {"imported": True})
        storage.save_lot(lot)
        created.append({"lot_id": lot_id, "producer": lot.producer_name, "n_plots": len(plots)})
    return created


TEMPLATE_CSV = (
    "productor,producto,region,lote,parcela,lat,lon,area_ha,cantidad_kg,poligono\n"
    "Ana Quispe,cafe,San Martin,LOTE-A,P1,-6.4781,-76.3720,2.5,900,\n"
    "Ana Quispe,cafe,San Martin,LOTE-A,P2,-6.4760,-76.3705,1.8,600,\n"
    "Beto Rios,cacao,Cusco,,P1,,,5.2,1800,\"-12.869,-72.941; -12.869,-72.938; -12.866,-72.938; -12.866,-72.941\"\n"
)


def template_xlsx_bytes():
    """La misma plantilla, en Excel (.xlsx) con encabezados destacados y notas de uso."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "Parcelas"
    rows = [r.split(",") for r in TEMPLATE_CSV.strip().splitlines()]
    # la fila del polígono trae comas entrecomilladas: re-parsear con csv
    import csv as _csv
    rows = list(_csv.reader(_io.StringIO(TEMPLATE_CSV)))
    for r in rows:
        ws.append(r)
    hf = Font(bold=True, color="FFFFFF")
    for c in ws[1]:
        c.font = hf; c.fill = PatternFill("solid", fgColor="0B3D2E")
    for col, w in zip("ABCDEFGHIJ", (16, 10, 12, 10, 9, 11, 11, 9, 12, 46)):
        ws.column_dimensions[col].width = w
    notes = wb.create_sheet("Instrucciones")
    for t in (
        "Cómo usar esta plantilla (una fila = una parcela):",
        "· productor: nombre del agricultor. · producto: cafe o cacao. · region: departamento/zona.",
        "· lote (opcional): filas con el mismo lote se agrupan en un solo lote. Si se deja vacío, se agrupa por productor.",
        "· lat / lon: coordenadas decimales con al menos 5 decimales (en Perú lat es negativa, ej. -6.478123).",
        "· area_ha: hectáreas de la parcela. · cantidad_kg: kilos estimados de la parcela.",
        "· poligono (opcional, para parcelas >4 ha): vértices 'lat,lon; lat,lon; ...' (mínimo 3). Si se llena, lat/lon pueden quedar vacíos.",
        "Guarde el archivo y súbalo en 'Importar parcelas' (app o panel).",
    ):
        notes.append([t])
    notes.column_dimensions["A"].width = 110
    buf = _io.BytesIO(); wb.save(buf)
    return buf.getvalue()
